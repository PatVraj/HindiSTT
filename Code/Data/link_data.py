import openpyxl

def read_file(filename):
    dates_links = []
    with open(filename, 'r') as file:
        lines = file.readlines()
        for i in range(0, len(lines), 2):
            try:
                date = lines[i].strip().split(': ')[1]
                link = lines[i + 1].strip().split(': ')[1]
                dates_links.append((date, link))
            except IndexError:
                pass
    return dates_links

def create_excel_file(dates_links, output_filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Dates and Links'
    
    # Write headers
    sheet['A1'] = 'Date'
    sheet['B1'] = 'Link'
    
    # Write data
    for row, (date, link) in enumerate(dates_links, start=2):
        sheet.cell(row=row, column=1, value=date)
        sheet.cell(row=row, column=2, value=link)
    
    workbook.save(output_filename)
    print(f'Excel file "{output_filename}" created successfully.')

# Input and output filenames
input_filename = 'speech_links.txt'
output_filename = 'speech_data.xlsx'

# Read the text file
dates_links = read_file(input_filename)
# Create the Excel file
create_excel_file(dates_links, output_filename)